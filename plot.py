def create_results_table(problem_sizes, normal_times, efficient_times, normal_memories, efficient_memories):
    """Create a formatted table with benchmark results"""
    # Generate markdown table
    markdown_table = "| Problem Size | Normal DP Time (ms) | Efficient DP Time (ms) | Normal DP Memory (KB) | Efficient DP Memory (KB) |\n"
    markdown_table += "|------------:|-------------------:|----------------------:|---------------------:|------------------------:|\n"
    
    for i in range(len(problem_sizes)):
        markdown_table += f"| {problem_sizes[i]:,} | {normal_times[i]:.5f} | {efficient_times[i]:.5f} | {normal_memories[i]:,} | {efficient_memories[i]:,} |\n"
    
    # Generate HTML table
    html_table = """
    <table border="1" cellpadding="5" cellspacing="0" style="border-collapse: collapse; width: 100%;">
        <tr style="background-color: #f2f2f2; font-weight: bold; text-align: right;">
            <th>Problem Size</th>
            <th>Normal DP Time (ms)</th>
            <th>Efficient DP Time (ms)</th>
            <th>Normal DP Memory (KB)</th>
            <th>Efficient DP Memory (KB)</th>
        </tr>
    """
    
    for i in range(len(problem_sizes)):
        html_table += f"""
        <tr style="text-align: right;">
            <td>{problem_sizes[i]:,}</td>
            <td>{normal_times[i]:.5f}</td>
            <td>{efficient_times[i]:.5f}</td>
            <td>{normal_memories[i]:,}</td>
            <td>{efficient_memories[i]:,}</td>
        </tr>
        """
    
    html_table += "</table>"
    
    # Generate TSV data (tab-separated values)
    tsv_data = "Problem Size\tNormal DP Time (ms)\tEfficient DP Time (ms)\tNormal DP Memory (KB)\tEfficient DP Memory (KB)\n"
    for i in range(len(problem_sizes)):
        tsv_data += f"{problem_sizes[i]}\t{normal_times[i]:.5f}\t{efficient_times[i]:.5f}\t{normal_memories[i]}\t{efficient_memories[i]}\n"
    
    # Generate plain text table (for terminal output)
    header = f"{'Problem Size':>15} | {'Normal Time (ms)':>20} | {'Efficient Time (ms)':>20} | {'Normal Memory (KB)':>20} | {'Efficient Memory (KB)':>20}"
    separator = "-" * len(header)
    
    plain_table = [header, separator]
    for i in range(len(problem_sizes)):
        row = f"{problem_sizes[i]:>15,} | {normal_times[i]:>20.5f} | {efficient_times[i]:>20.5f} | {normal_memories[i]:>20,} | {efficient_memories[i]:>20,}"
        plain_table.append(row)
    
    return markdown_table, html_table, tsv_data, "\n".join(plain_table)#!/usr/bin/env python3
import os
import sys
import subprocess
import matplotlib.pyplot as plt
import numpy as np
import argparse
import csv
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# File paths for the two implementations
NORMAL_DP_SCRIPT = "basic_3.py"  # Assuming this is your base file
EFFICIENT_DP_SCRIPT = "efficient_3.py"  # Your memory efficient implementation

# Plotting options
USE_LOG_SCALE_TIME = True    # Set to True to allow automatic log scaling for time plot
USE_LOG_SCALE_MEMORY = False  # Set to False to disable log scaling for memory plot

def get_file_size(file_path):
    """Get the size of input problem by reading the file and calculating final sequence lengths"""
    # Use the read_data function logic to calculate actual sequence sizes
    s1, s2 = read_data(file_path)
    return len(s1) + len(s2)

def read_data(input_file):
    """Read and expand the sequences according to the pattern in the input file"""
    with open(input_file, 'r') as f:
        s1, s2 = "", ""
        for line in f:
            line = line.strip()
            if line.isnumeric():
                l = int(line)
                if s2 == "":
                    s1 = s1[:l+1] + s1 + s1[l+1:]
                else:
                    s2 = s2[:l+1] + s2 + s2[l+1:]
            else:
                if s1 == "":
                    s1 = line
                else:
                    s2 = line
    return s1, s2

def run_algorithm(script_path, input_file, output_file):
    """Run the specified algorithm script on the input file and return time and memory usage"""
    result = subprocess.run(
        [sys.executable, script_path, input_file, output_file],
        capture_output=True,
        text=True
    )
    
    # Read time and memory from output file
    try:
        with open(output_file, 'r') as f:
            lines = f.readlines()
            # Based on your file format where the last two lines are time and memory
            if len(lines) >= 5:  # Accounting for cost value and aligned sequences
                time_used = float(lines[-2].strip())
                memory_used = int(lines[-1].strip())
                return time_used, memory_used
    except Exception as e:
        print(f"Error reading output file: {e}")
    
    # Return default values if we couldn't read the output properly
    return 0, 0

def main():
    # Parse command-line arguments
    parser = argparse.ArgumentParser(description='Benchmark DP algorithms and generate plots.')
    parser.add_argument('input_dir', help='Directory containing input files')
    parser.add_argument('output_dir', help='Directory for output files and plots')
    parser.add_argument('--normal-script', default=NORMAL_DP_SCRIPT, help=f'Path to normal DP script (default: {NORMAL_DP_SCRIPT})')
    parser.add_argument('--efficient-script', default=EFFICIENT_DP_SCRIPT, help=f'Path to memory-efficient DP script (default: {EFFICIENT_DP_SCRIPT})')
    parser.add_argument('--log-time', action='store_true', default=USE_LOG_SCALE_TIME, help='Use logarithmic scale for time plot if appropriate')
    parser.add_argument('--log-memory', action='store_true', default=USE_LOG_SCALE_MEMORY, help='Use logarithmic scale for memory plot if appropriate')
    parser.add_argument('--no-log-time', action='store_false', dest='log_time', help='Disable logarithmic scale for time plot')
    parser.add_argument('--no-log-memory', action='store_false', dest='log_memory', help='Disable logarithmic scale for memory plot')
    parser.add_argument('--table-format', choices=['all', 'md', 'html', 'tsv', 'csv', 'excel'], default='all', 
                        help='Format(s) for the results table (default: all)')
    parser.add_argument('--no-plots', action='store_true', help='Skip generating plots (generate only tables)')
    parser.add_argument('--sort', choices=['asc', 'desc', 'none'], default='asc',
                        help='Sort results by problem size (asc=ascending, desc=descending, none=as processed)')
    parser.add_argument('--excel-file', default='results_table.xlsx', help='Filename for Excel output (default: results_table.xlsx)')
    parser.add_argument('--csv-file', default='results_table.csv', help='Filename for CSV output (default: results_table.csv)')
    
    args = parser.parse_args()
    
    input_dir = args.input_dir
    output_dir = args.output_dir
    normal_script = args.normal_script
    efficient_script = args.efficient_script
    use_log_time = args.log_time
    use_log_memory = args.log_memory
    table_format = args.table_format
    generate_plots = not args.no_plots
    sort_order = args.sort
    excel_filename = args.excel_file
    csv_filename = args.csv_file
    
    # Check if script files exist
    if not os.path.isfile(normal_script):
        print(f"Error: Normal DP script '{normal_script}' not found.")
        sys.exit(1)
    
    if not os.path.isfile(efficient_script):
        print(f"Error: Efficient DP script '{efficient_script}' not found.")
        sys.exit(1)
    
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Data structures to store results
    problem_sizes = []
    normal_times = []
    normal_memories = []
    efficient_times = []
    efficient_memories = []
    
    # Process each input file
    input_files = [f for f in os.listdir(input_dir) if os.path.isfile(os.path.join(input_dir, f))]
    
    if not input_files:
        print(f"No input files found in directory: {input_dir}")
        sys.exit(1)
    
    print(f"Found {len(input_files)} input files to process.")
    
    for idx, input_file in enumerate(input_files, 1):
        input_path = os.path.join(input_dir, input_file)
        try:
            print(f"\nProcessing file {idx}/{len(input_files)}: {input_file}")
            problem_size = get_file_size(input_path)
            print(f"Calculated problem size: {problem_size} characters")
            problem_sizes.append(problem_size)
            
            # Run normal DP algorithm
            print("Running normal DP algorithm...")
            normal_output = os.path.join(output_dir, f"normal_{input_file}_output.txt")
            normal_time, normal_memory = run_algorithm(normal_script, input_path, normal_output)
            normal_times.append(normal_time)
            normal_memories.append(normal_memory)
            
            # Run memory-efficient algorithm
            print("Running memory-efficient DP algorithm...")
            efficient_output = os.path.join(output_dir, f"efficient_{input_file}_output.txt")
            efficient_time, efficient_memory = run_algorithm(efficient_script, input_path, efficient_output)
            efficient_times.append(efficient_time)
            efficient_memories.append(efficient_memory)
            
            print(f"Results for {input_file} (size: {problem_size}):")
            print(f"  Normal DP: Time = {normal_time:.5f}ms, Memory = {normal_memory} KB")
            print(f"  Efficient: Time = {efficient_time:.5f}ms, Memory = {efficient_memory} KB")
            
            # Print improvement metrics
            time_improvement = (normal_time - efficient_time) / normal_time * 100 if normal_time > 0 else 0
            memory_improvement = (normal_memory - efficient_memory) / normal_memory * 100 if normal_memory > 0 else 0
            print(f"  Improvements: Time = {time_improvement:.5f}%, Memory = {memory_improvement:.5f}%")
            
        except Exception as e:
            print(f"Error processing file {input_file}: {e}")
            # Continue with next file instead of stopping completely
            continue
    
    # Filter out any data points with zero problem size or missing metrics
    valid_indices = []
    for i in range(len(problem_sizes)):
        if (problem_sizes[i] > 0 and 
            normal_times[i] > 0 and efficient_times[i] > 0 and 
            normal_memories[i] > 0 and efficient_memories[i] > 0):
            valid_indices.append(i)
    
    if not valid_indices:
        print("No valid data points to plot. Check if algorithms ran successfully.")
        sys.exit(1)
    
    problem_sizes = [problem_sizes[i] for i in valid_indices]
    normal_times = [normal_times[i] for i in valid_indices]
    normal_memories = [normal_memories[i] for i in valid_indices]
    efficient_times = [efficient_times[i] for i in valid_indices]
    efficient_memories = [efficient_memories[i] for i in valid_indices]
    
    # Sort data points by problem size if requested
    if sort_order != 'none':
        sorted_indices = np.argsort(problem_sizes)
        if sort_order == 'desc':
            sorted_indices = sorted_indices[::-1]  # Reverse for descending order
            
        problem_sizes = [problem_sizes[i] for i in sorted_indices]
        normal_times = [normal_times[i] for i in sorted_indices]
        normal_memories = [normal_memories[i] for i in sorted_indices]
        efficient_times = [efficient_times[i] for i in sorted_indices]
        efficient_memories = [efficient_memories[i] for i in sorted_indices]
    
    print(f"\nGenerating results for {len(problem_sizes)} valid data points.")
    
    # Generate plots if requested
    time_plot_path = None
    memory_plot_path = None
    
    if generate_plots:
        # Generate time vs problem size plot
        plt.figure(figsize=(12, 7))
        plt.plot(problem_sizes, normal_times, 'o-', color='blue', label='Normal DP')
        plt.plot(problem_sizes, efficient_times, 's-', color='red', label='Memory-Efficient DP')
        plt.xlabel('Problem Size (Total Sequence Length)', fontsize=12)
        plt.ylabel('Execution Time (ms)', fontsize=12)
        plt.title('Execution Time vs Problem Size', fontsize=14)
        plt.legend(fontsize=10)
        plt.grid(True, linestyle='--', alpha=0.7)
        
        # Add data point annotations
        for i, (size, time1, time2) in enumerate(zip(problem_sizes, normal_times, efficient_times)):
            plt.annotate(f'{size}', (size, time1), textcoords="offset points", 
                        xytext=(0,10), ha='center', fontsize=8)
        
        # Use logarithmic scale if enabled and range is large (more than 10x difference)
        if use_log_time and max(normal_times + efficient_times) / max(1, min(normal_times + efficient_times)) > 10:
            plt.yscale('log')
            plt.ylabel('Execution Time (ms) - Log Scale', fontsize=12)
        
        plt.tight_layout()
        time_plot_path = os.path.join(output_dir, 'time_comparison.png')
        plt.savefig(time_plot_path, dpi=300)
        
        # Generate memory vs problem size plot
        plt.figure(figsize=(12, 7))
        plt.plot(problem_sizes, normal_memories, 'o-', color='blue', label='Normal DP')
        plt.plot(problem_sizes, efficient_memories, 's-', color='red', label='Memory-Efficient DP')
        plt.xlabel('Problem Size (Total Sequence Length)', fontsize=12)
        plt.ylabel('Memory Usage (KB)', fontsize=12)
        plt.title('Memory Usage vs Problem Size', fontsize=14)
        plt.legend(fontsize=10)
        plt.grid(True, linestyle='--', alpha=0.7)
        
        # Add data point annotations
        for i, (size, mem1, mem2) in enumerate(zip(problem_sizes, normal_memories, efficient_memories)):
            plt.annotate(f'{size}', (size, mem1), textcoords="offset points", 
                        xytext=(0,10), ha='center', fontsize=8)
        
        # Use logarithmic scale if enabled and range is large (more than 10x difference)
        if use_log_memory and max(normal_memories + efficient_memories) / max(1, min(normal_memories + efficient_memories)) > 10:
            plt.yscale('log')
            plt.ylabel('Memory Usage (KB) - Log Scale', fontsize=12)
        
        plt.tight_layout()
        memory_plot_path = os.path.join(output_dir, 'memory_comparison.png')
        plt.savefig(memory_plot_path, dpi=300)
    
    # Save the raw data as CSV for further analysis
    csv_path = os.path.join(output_dir, 'benchmark_results.csv')
    with open(csv_path, 'w') as f:
        f.write('problem_size,normal_time_ms,normal_memory_kb,efficient_time_ms,efficient_memory_kb,time_improvement_pct,memory_improvement_pct\n')
        for i in range(len(problem_sizes)):
            time_imp = (normal_times[i] - efficient_times[i]) / normal_times[i] * 100 if normal_times[i] > 0 else 0
            mem_imp = (normal_memories[i] - efficient_memories[i]) / normal_memories[i] * 100 if normal_memories[i] > 0 else 0
            f.write(f"{problem_sizes[i]},{normal_times[i]:.5f},{normal_memories[i]},{efficient_times[i]:.5f},{efficient_memories[i]},{time_imp:.5f},{mem_imp:.5f}\n")
    
    # Create a summary for average improvements
    avg_time_imp = sum([(normal_times[i] - efficient_times[i]) / normal_times[i] * 100 if normal_times[i] > 0 else 0 
                         for i in range(len(problem_sizes))]) / len(problem_sizes)
    avg_mem_imp = sum([(normal_memories[i] - efficient_memories[i]) / normal_memories[i] * 100 if normal_memories[i] > 0 else 0 
                        for i in range(len(problem_sizes))]) / len(problem_sizes)
    
    summary_path = os.path.join(output_dir, 'summary.txt')
    with open(summary_path, 'w') as f:
        f.write(f"Benchmark Summary\n")
        f.write(f"----------------\n\n")
        f.write(f"Number of test cases: {len(problem_sizes)}\n")
        f.write(f"Problem size range: {min(problem_sizes)} to {max(problem_sizes)} characters\n\n")
        f.write(f"Average time improvement: {avg_time_imp:.5f}%\n")
        f.write(f"Average memory improvement: {avg_mem_imp:.5f}%\n\n")
        f.write(f"Best time improvement: {max([(normal_times[i] - efficient_times[i]) / normal_times[i] * 100 if normal_times[i] > 0 else 0 for i in range(len(problem_sizes))]):.5f}%\n")
        f.write(f"Best memory improvement: {max([(normal_memories[i] - efficient_memories[i]) / normal_memories[i] * 100 if normal_memories[i] > 0 else 0 for i in range(len(problem_sizes))]):.5f}%\n")
    
    # Create formatted tables of results
    markdown_table, html_table, tsv_data, plain_table = create_results_table(
        problem_sizes, normal_times, efficient_times, normal_memories, efficient_memories
    )
    
    # Save tables in requested format(s)
    md_table_path = html_table_path = tsv_table_path = csv_table_path = excel_table_path = None
    table_paths = []
    
    if table_format in ['all', 'md']:
        md_table_path = os.path.join(output_dir, 'results_table.md')
        with open(md_table_path, 'w') as f:
            f.write("# Benchmark Results Table\n\n")
            f.write(markdown_table)
        table_paths.append(f"  * Markdown: {md_table_path}")
    
    if table_format in ['all', 'html']:
        html_table_path = os.path.join(output_dir, 'results_table.html')
        with open(html_table_path, 'w') as f:
            f.write("<html><head><title>Benchmark Results</title></head><body>\n")
            f.write("<h1>Benchmark Results Table</h1>\n")
            f.write(html_table)
            f.write("\n</body></html>")
        table_paths.append(f"  * HTML: {html_table_path}")
    
    if table_format in ['all', 'tsv']:
        tsv_table_path = os.path.join(output_dir, 'results_table.tsv')
        with open(tsv_table_path, 'w') as f:
            f.write(tsv_data)
        table_paths.append(f"  * TSV: {tsv_table_path}")
    
    if table_format in ['all', 'csv']:
        csv_table_path = os.path.join(output_dir, csv_filename)
        with open(csv_table_path, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(["Problem Size", "Normal DP Time (ms)", "Efficient DP Time (ms)", 
                            "Normal DP Memory (KB)", "Efficient DP Memory (KB)"])
            for i in range(len(problem_sizes)):
                writer.writerow([problem_sizes[i], f"{normal_times[i]:.5f}", f"{efficient_times[i]:.5f}", 
                                normal_memories[i], efficient_memories[i]])
        table_paths.append(f"  * CSV: {csv_table_path}")
    
    if table_format in ['all', 'excel'] and EXCEL_AVAILABLE:
        excel_table_path = os.path.join(output_dir, excel_filename)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Benchmark Results"
        
        # Add headers with formatting
        headers = ["Problem Size", "Normal DP Time (ms)", "Efficient DP Time (ms)", 
                  "Normal DP Memory (KB)", "Efficient DP Memory (KB)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # Add data
        for row, size in enumerate(problem_sizes, 2):
            ws.cell(row=row, column=1, value=size)
            ws.cell(row=row, column=2, value=round(normal_times[row-2], 2))
            ws.cell(row=row, column=3, value=round(efficient_times[row-2], 2))
            ws.cell(row=row, column=4, value=normal_memories[row-2])
            ws.cell(row=row, column=5, value=efficient_memories[row-2])
        
        # Auto-adjust column widths
        for col in range(1, 6):
            max_length = 0
            for row in range(1, len(problem_sizes) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_length + 4
        
        # Add a second sheet with improvement percentages
        ws_imp = wb.create_sheet(title="Improvements")
        
        # Add headers
        imp_headers = ["Problem Size", "Time Improvement (%)", "Memory Improvement (%)"]
        for col, header in enumerate(imp_headers, 1):
            cell = ws_imp.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # Add data
        for row, size in enumerate(problem_sizes, 2):
            time_imp = (normal_times[row-2] - efficient_times[row-2]) / normal_times[row-2] * 100 if normal_times[row-2] > 0 else 0
            mem_imp = (normal_memories[row-2] - efficient_memories[row-2]) / normal_memories[row-2] * 100 if normal_memories[row-2] > 0 else 0
            
            ws_imp.cell(row=row, column=1, value=size)
            ws_imp.cell(row=row, column=2, value=round(time_imp, 2))
            ws_imp.cell(row=row, column=3, value=round(mem_imp, 2))
        
        # Auto-adjust column widths
        for col in range(1, 4):
            max_length = 0
            for row in range(1, len(problem_sizes) + 2):
                cell_value = ws_imp.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws_imp.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_length + 4
        
        wb.save(excel_table_path)
        table_paths.append(f"  * Excel: {excel_table_path}")
    elif table_format in ['all', 'excel'] and not EXCEL_AVAILABLE:
        print("Warning: Excel format requested but openpyxl package is not installed.")
        print("To enable Excel output, install it with: pip install openpyxl")
        table_paths.append(f"  * Excel: Not available (openpyxl not installed)")
    
    
    # Print the table to console
    print("\nResults Table:")
    print(plain_table)
    
    # Add table to the summary file
    with open(summary_path, 'a') as f:
        f.write("\n\n## Results Table\n\n")
        f.write(markdown_table)
    
    print(f"\nResults saved to {output_dir}")
    if generate_plots and time_plot_path and memory_plot_path:
        print(f"- Time comparison plot: {time_plot_path}")
        print(f"- Memory comparison plot: {memory_plot_path}")
    if table_paths:
        print(f"- Results tables:")
        for path in table_paths:
            print(path)
    print(f"- Raw data: {csv_path}")
    print(f"- Summary report: {summary_path}")
    print("\nSummary:")
    print(f"- Average time improvement: {avg_time_imp:.5f}%")
    print(f"- Average memory improvement: {avg_mem_imp:.5f}% Markdown: {md_table_path}")
    print(f"  * HTML: {html_table_path}")
    print(f"  * TSV: {tsv_table_path}")
    print(f"- Raw data: {csv_path}")
    print(f"- Summary report: {summary_path}")
    print("\nSummary:")
    print(f"- Average time improvement: {avg_time_imp:.5f}%")
    print(f"- Average memory improvement: {avg_mem_imp:.5f}%")

if __name__ == "__main__":
    main()